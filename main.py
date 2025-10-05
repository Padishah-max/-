# main.py — режим WEBHOOK для Render, админ запускает викторину
import os, json, sqlite3, asyncio, time, logging
from dataclasses import dataclass
from typing import List, Optional, Dict, Set, Tuple

import httpx
import openpyxl
from openpyxl.utils import get_column_letter

from telegram import Update, InlineKeyboardMarkup, InlineKeyboardButton
from telegram.ext import (
    ApplicationBuilder, Application, ContextTypes,
    CommandHandler, CallbackQueryHandler, PollAnswerHandler
)

# ---------- ЛОГИ ----------
logging.basicConfig(level=logging.INFO)
log = logging.getLogger("quizbot")

# ---------- ОКРУЖЕНИЕ ----------
BOT_TOKEN   = (os.getenv("BOT_TOKEN")   or "").strip()
PUBLIC_URL  = (os.getenv("PUBLIC_URL")  or "").strip().rstrip("/")
PORT        = int(os.getenv("PORT", "10000"))
ADMINS_ENV  = os.getenv("ADMINS", "")

if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN is empty (Settings → Environment).")
if not PUBLIC_URL:
    raise RuntimeError("PUBLIC_URL is empty (например, https://your.onrender.com).")

ADMIN_IDS = {int(x) for x in ADMINS_ENV.split(",") if x.strip().isdigit()} or {133637780}

DB_FILE        = "quiz.db"
QUESTIONS_FILE = "questions.json"
COUNTRIES      = ["Россия", "Казахстан", "Армения", "Беларусь", "Кыргызстан"]
QUESTION_SECONDS = 30
COUNTDOWN         = 3

# Глобальный флаг «идёт раунд»
QUIZ_ACTIVE = False

# ---------- МОДЕЛИ ----------
@dataclass
class Question:
    text: str
    options: List[str]
    correct: List[int]
    multiple: bool

@dataclass
class UserQuizState:
    index: int = 0
    last_poll_id: Optional[str] = None
    started: bool = False
    finished: bool = False

QUESTIONS: List[Question] = []
STATE: Dict[int, UserQuizState] = {}

# ---------- БАЗА ----------
def db():
    conn = sqlite3.connect(DB_FILE)
    conn.execute("""CREATE TABLE IF NOT EXISTS users(
        user_id INTEGER PRIMARY KEY,
        country TEXT
    )""")
    conn.execute("""CREATE TABLE IF NOT EXISTS answers(
        user_id INTEGER,
        q_index INTEGER,
        option_ids TEXT,
        correct INTEGER
    )""")
    return conn

def is_admin(uid: int) -> bool:
    return uid in ADMIN_IDS

def get_registered_users() -> List[int]:
    with db() as conn:
        rows = conn.execute("SELECT user_id FROM users WHERE country IS NOT NULL").fetchall()
    return [r[0] for r in rows]

# ---------- ВОПРОСЫ ----------
def _validate_questions(data: List[dict]) -> List[Question]:
    out: List[Question] = []
    if not isinstance(data, list):
        raise ValueError("Root of questions JSON must be a list.")
    for i, q in enumerate(data, start=1):
        text = (q.get("text") or "").strip()
        options = q.get("options") or []
        correct = q.get("correct_indices") or []
        multiple = bool(q.get("multiple", False))
        if not text or not isinstance(options, list) or len(options) < 2:
            raise ValueError(f"Q{i}: invalid text/options")
        if not isinstance(correct, list) or not all(isinstance(ci, int) for ci in correct):
            raise ValueError(f"Q{i}: correct_indices must be list[int]")
        if any(ci < 0 or ci >= len(options) for ci in correct):
            raise ValueError(f"Q{i}: correct index out of range")
        if not correct:
            raise ValueError(f"Q{i}: at least one correct answer required")
        if len(correct) > 1 and not multiple:
            raise ValueError(f"Q{i}: multiple answers but multiple=false")
        out.append(Question(text, options, correct, multiple))
    return out

def load_questions_from_file() -> int:
    global QUESTIONS
    if not os.path.exists(QUESTIONS_FILE):
        QUESTIONS = []
        log.warning("%s not found. No questions loaded.", QUESTIONS_FILE)
        return 0
    with open(QUESTIONS_FILE, "r", encoding="utf-8") as f:
        data = json.load(f)
    QUESTIONS = _validate_questions(data)
    log.info("Loaded %d questions.", len(QUESTIONS))
    return len(QUESTIONS)

async def set_questions_from_url(url: str) -> int:
    async with httpx.AsyncClient(timeout=30) as client:
        r = await client.get(url)
        r.raise_for_status()
        data = r.json()
    _ = _validate_questions(data)
    with open(QUESTIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return load_questions_from_file()

# ---------- СОСТОЯНИЕ ----------
def st(uid: int) -> UserQuizState:
    if uid not in STATE:
        STATE[uid] = UserQuizState()
    return STATE[uid]

def reset_user(uid: int):
    STATE.pop(uid, None)
    with db() as conn:
        conn.execute("DELETE FROM answers WHERE user_id=?", (uid,))

# ---------- КВИЗ (личный чат) ----------
async def start_user_quiz(uid: int, ctx: ContextTypes.DEFAULT_TYPE, countdown: int = COUNTDOWN):
    """Запустить попытку у конкретного пользователя (если активен раунд)."""
    s = st(uid)
    if s.started and not s.finished:
        return
    s.index = 0
    s.finished = False
    s.started = True

    if countdown > 0:
        msg = await ctx.bot.send_message(uid, f"🚀 Старт через {countdown}…")
        for left in range(countdown - 1, 0, -1):
            await asyncio.sleep(1)
            try:
                await msg.edit_text(f"🚀 Старт через {left}…")
            except:
                pass
        try:
            await msg.edit_text("🔥 Поехали!")
        except:
            await ctx.bot.send_message(uid, "🔥 Поехали!")

    await send_next(uid, ctx)

async def send_next(uid: int, ctx: ContextTypes.DEFAULT_TYPE):
    s = st(uid)
    if s.index >= len(QUESTIONS):
        s.finished = True
        await ctx.bot.send_message(uid, "✅ Спасибо! Ваши ответы сохранены.")
        return

    q = QUESTIONS[s.index]
    head = f"Вопрос {s.index+1}/{len(QUESTIONS)}"
    suffix = " (несколько ответов)" if q.multiple else ""
    rules = f"\n⏱ На ответ даётся {QUESTION_SECONDS} секунд."
    text = f"{head}\n{q.text}{suffix}{rules}"

    msg = await ctx.bot.send_poll(
        chat_id=uid,
        question=text,
        options=q.options,
        is_anonymous=False,
        allows_multiple_answers=q.multiple
    )
    s.last_poll_id = msg.poll.id

    async def q_timeout(poll_id: str, user_id: int):
        await asyncio.sleep(QUESTION_SECONDS)
        s2 = st(user_id)
        if s2.last_poll_id == poll_id and not s2.finished:
            await ctx.bot.send_message(user_id, "⏰ Время на этот вопрос вышло.")
            s2.index += 1
            await send_next(user_id, ctx)

    asyncio.create_task(q_timeout(msg.poll.id, uid))

# ---------- ОТЧЁТ ----------
async def export_results_file() -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Answers"
    ws.append(["Страна","Пользователь","Вопрос №","Ответ(ы) индексы","Правильно"])

    answers: List[Tuple[int,int,str,int]] = []
    user_country: Dict[int,str] = {}
    with db() as conn:
        for uid,cnt in conn.execute("SELECT user_id,country FROM users"):
            user_country[uid] = cnt or "?"
        for row in conn.execute("SELECT user_id,q_index,option_ids,correct FROM answers"):
            answers.append(row)

    for uid,qidx,opt,corr in answers:
        ws.append([user_country.get(uid,"?"), uid, qidx+1, opt, "Да" if corr else "Нет"])
    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 18

    ws2 = wb.create_sheet("ByCountry")
    ws2.append(["Страна","Участников","Ответов всего","Правильных","Неправильных","Точность, %"])
    stats: Dict[str, Dict[str,int]] = {}
    users: Dict[str, Set[int]] = {}
    for uid,qidx,opt,corr in answers:
        c = user_country.get(uid,"?")
        stats.setdefault(c, {"tot":0,"ok":0})
        users.setdefault(c,set()).add(uid)
        stats[c]["tot"] += 1
        if corr: stats[c]["ok"] += 1
    for c,s in stats.items():
        tot=s["tot"]; ok=s["ok"]; bad=tot-ok; ppl=len(users.get(c,set()))
        acc=round((ok/tot*100) if tot else 0.0,2)
        ws2.append([c,ppl,tot,ok,bad,acc])
    for col in ws2.columns:
        ws2.column_dimensions[get_column_letter(col[0].column)].width = 18

    ws3 = wb.create_sheet("ByCountryQuestion")
    ws3.append(["Страна","Вопрос №","Ответов","Правильных","Неправильных","Точность, %"])
    cqm: Dict[Tuple[str,int],Dict[str,int]] = {}
    for uid,qidx,opt,corr in answers:
        c=user_country.get(uid,"?")
        key=(c,qidx+1)
        d=cqm.setdefault(key,{"tot":0,"ok":0})
        d["tot"]+=1
        if corr: d["ok"]+=1
    for (c,qn),d in sorted(cqm.items(), key=lambda x:(x[0][0],x[0][1])):
        tot=d["tot"]; ok=d["ok"]; bad=tot-ok; acc=round((ok/tot*100) if tot else 0.0,2)
        ws3.append([c,qn,tot,ok,bad,acc])
    for col in ws3.columns:
        ws3.column_dimensions[get_column_letter(col[0].column)].width = 18

    path = f"results_{int(time.time())}.xlsx"
    wb.save(path)
    return path

# ---------- КОМАНДЫ (участник) ----------
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Участник выбирает страну и ждёт старта от админа."""
    if update.effective_chat.type != "private":
        await update.message.reply_text("Напишите мне в личном чате, чтобы пройти викторину.")
        return
    kb = [[InlineKeyboardButton(c, callback_data=f"set_country:{c}") ] for c in COUNTRIES]
    await update.message.reply_text(
        "Выберите страну. После выбора вы будете зарегистрированы и увидите сообщение: "
        "«Ожидайте старт от организатора». Админ запустит викторину для всех зарегистрированных.",
        reply_markup=InlineKeyboardMarkup(kb)
    )

async def cmd_again(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Сброс своей попытки; ждём следующего админ-старта."""
    uid = update.effective_user.id
    reset_user(uid)
    kb = [[InlineKeyboardButton(c, callback_data=f"set_country:{c}") ] for c in COUNTRIES]
    await update.message.reply_text(
        "🔁 Пройти заново. Выберите страну — и ожидайте старт от организатора.",
        reply_markup=InlineKeyboardMarkup(kb)
    )

async def cmd_help(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🤖 Личная викторина.\n"
        "/start — регистрация (выбор страны) и ожидание старта\n"
        "/again — сброс своей попытки\n\n"
        "Админ:\n"
        "/start_quiz — запустить викторину для всех зарегистрированных\n"
        "/report — Excel-отчёт\n"
        "/reload — перечитать questions.json\n"
        "/setq <raw_json_url> — загрузить вопросы по URL\n"
        "/status — зарегистрированные пользователи по странам"
    )

# ---------- КОМАНДЫ (админ) ----------
async def cmd_start_quiz(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """АДМИН: запустить для всех зарегистрированных, кто ещё не начал или уже сбросил."""
    global QUIZ_ACTIVE
    if not is_admin(update.effective_user.id):
        return
    if not QUESTIONS:
        await update.message.reply_text("Нет загруженных вопросов. Используй /reload или /setq.")
        return

    QUIZ_ACTIVE = True
    uids = get_registered_users()
    await update.message.reply_text(f"▶️ Запускаю викторину для {len(uids)} зарегистрированных пользователей.")

    # Стартуем всем, у кого нет активной попытки
    started = 0
    for uid in uids:
        s = st(uid)
        if not s.started or s.finished:
            try:
                await update.get_bot().send_message(uid, "Организатор запустил викторину.")
                await start_user_quiz(uid, ctx, COUNTDOWN)
                started += 1
            except Exception as e:
                log.warning("Cannot start for %s: %s", uid, e)

    await update.message.reply_text(f"Готово. Запущено попыток: {started}/{len(uids)}.")

async def cmd_report(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    await update.message.reply_text("Формирую отчёт…")
    path = await export_results_file()
    await ctx.bot.send_document(update.effective_user.id, open(path, "rb"), filename=os.path.basename(path))

async def cmd_reload(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    n = load_questions_from_file()
    await update.message.reply_text(f"Перечитал {QUESTIONS_FILE}: вопросов {n}.")

async def cmd_setq(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    if not ctx.args:
        await update.message.reply_text("Укажи RAW-ссылку: /setq https://raw.githubusercontent.com/.../questions.json")
        return
    url = ctx.args[0]
    try:
        n = await set_questions_from_url(url)
        await update.message.reply_text(f"Загрузил новые вопросы. Всего: {n}.")
    except Exception as e:
        await update.message.reply_text(f"Ошибка: {e}")

async def cmd_status(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update.effective_user.id):
        return
    with db() as conn:
        rows = conn.execute("SELECT country, COUNT(*) FROM users WHERE country IS NOT NULL GROUP BY country").fetchall()
    total = sum(r[1] for r in rows)
    lines = [f"Всего зарегистрировано: {total}"]
    for c, cnt in rows:
        lines.append(f"- {c}: {cnt}")
    await update.message.reply_text("\n".join(lines))

# ---------- КНОПКИ ----------
async def on_button(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    cq = update.callback_query
    data = cq.data or ""

    if data.startswith("set_country:"):
        country = data.split(":",1)[1]
        uid = cq.from_user.id

        # Запишем/обновим страну (не стартуем автоматически!)
        with db() as conn:
            conn.execute(
                "INSERT INTO users(user_id,country) VALUES(?,?) "
                "ON CONFLICT(user_id) DO UPDATE SET country=excluded.country",
                (uid, country)
            )

        # Сбрасывать ответы НЕ будем — человек мог выбрать страну повторно.
        # Если хочет заново — у него есть /again.
        msg = (
            f"Страна: {country} сохранена.\n"
            f"Ожидайте старт от организатора. Время на каждый вопрос — {QUESTION_SECONDS} сек.\n"
            f"Чтобы пройти заново позже: воспользуйтесь командой /again."
        )
        try:
            await cq.edit_message_text(msg)
        except:
            await cq.message.reply_text(msg)

# ---------- POLL ----------
async def on_poll_answer(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    ans = update.poll_answer
    uid = ans.user.id
    s = st(uid)
    if s.finished or s.last_poll_id != ans.poll_id:
        return

    q = QUESTIONS[s.index]
    chosen = ans.option_ids or []
    correct = int(set(chosen) == set(q.correct))
    with db() as conn:
        conn.execute(
            "INSERT INTO answers(user_id,q_index,option_ids,correct) VALUES(?,?,?,?)",
            (uid, s.index, json.dumps(chosen, ensure_ascii=False), correct)
        )
    s.index += 1
    await send_next(uid, ctx)

# ---------- APP ----------
def build_app() -> Application:
    load_questions_from_file()
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    # Участник
    app.add_handler(CommandHandler("start",  cmd_start))
    app.add_handler(CommandHandler("again",  cmd_again))
    app.add_handler(CommandHandler("help",   cmd_help))

    # Админ
    app.add_handler(CommandHandler("start_quiz", cmd_start_quiz))
    app.add_handler(CommandHandler("report",     cmd_report))
    app.add_handler(CommandHandler("reload",     cmd_reload))
    app.add_handler(CommandHandler("setq",       cmd_setq))
    app.add_handler(CommandHandler("status",     cmd_status))

    # Кнопки + ответы на опросы
    app.add_handler(CallbackQueryHandler(on_button))
    app.add_handler(PollAnswerHandler(on_poll_answer))
    return app

if __name__ == "__main__":
    application = build_app()
    log.info("Starting in WEBHOOK mode on port %s; PUBLIC_URL=%s", PORT, PUBLIC_URL)
    application.run_webhook(
        listen="0.0.0.0",
        port=PORT,
        url_path=BOT_TOKEN,                         # скрытый путь
        webhook_url=f"{PUBLIC_URL}/{BOT_TOKEN}",    # Telegram будет слать сюда
        drop_pending_updates=True,
        allowed_updates=["message","callback_query","poll_answer"]
    )
